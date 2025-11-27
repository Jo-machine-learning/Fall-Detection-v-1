import cv2
import cvzone
import math
from ultralytics import YOLO
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

source = 'vid3.mp4'
is_image = os.path.splitext(source)[1].lower() in ['.jpg', '.jpeg', '.png']

model = YOLO("yolov8n.pt") 
classnames = [line.strip() for line in open('classes.txt')]

excel_file = "fall_log.xlsx"


if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Datetime", "Frame", "Class", "Confidence"])
    wb.save(excel_file)

def log_fall(frame_number, class_name, confidence):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), frame_number, class_name, confidence])
    wb.save(excel_file)

def check_fall(width, height):
    """نسبة عرض/ارتفاع لتحديد السقوط"""
    return (height - width) < 0

fall_state = False
fall_color = (0, 0, 255)
animation_counter = 0

def update_fall_animation(detected):
    global fall_color, animation_counter
    if detected:
        fall_color = (0, 255, 0)
        animation_counter = 5
    elif animation_counter > 0:
        animation_counter -= 1
    else:
        fall_color = (0, 0, 255)

def process_frame(frame, frame_number):
    global fall_state
    frame = cv2.resize(frame, (640, 480))
    results = model(frame)
    fall_detected = False

    for info in results:
        for box in info.boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            confidence = float(box.conf[0])
            class_detect = int(box.cls[0])
            class_detect = classnames[class_detect]
            
            width = x2 - x1
            height = y2 - y1

            if confidence > 0.8 and class_detect == 'person':
                cvzone.cornerRect(frame, [x1, y1, width, height], l=20, rt=6, colorR=(255,0,0))
                
                if check_fall(width, height):
                    fall_detected = True
                    
                    log_fall(frame_number, class_detect, round(confidence, 2))

    update_fall_animation(fall_detected)
    
    
    cv2.rectangle(frame, (10,10), (180,60), fall_color, 2)
    text = "Fall Detected" if fall_detected else "No Fall"
    cv2.putText(frame, text, (20, 45), cv2.FONT_HERSHEY_SIMPLEX, 1, fall_color, 2)

    return frame

if is_image:
    frame = cv2.imread(source)
    frame = process_frame(frame, 1)
    cv2.imshow("Fall Detection", frame)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
else:
    cap = cv2.VideoCapture(source)
    fps = cap.get(cv2.CAP_PROP_FPS)
    delay = int(1000 / fps)
    frame_number = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        frame_number += 1
        frame = process_frame(frame, frame_number)
        cv2.imshow("Fall Detection", frame)
        if cv2.waitKey(delay) & 0xFF == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()
